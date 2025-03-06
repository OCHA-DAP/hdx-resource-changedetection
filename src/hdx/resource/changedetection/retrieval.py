"""Utility to download and hash resources. Uses asyncio. Note that the purpose of
asyncio is to help with IO-bound rather than CPU-bound code (for which multiprocessing
is more suitable as it leverages multiple CPUs). Asyncio allows you to structure your
code so that when one piece of linear single-threaded code (coroutine) is waiting for
something to happen another can take over and use the CPU. While conceptually similar to
threading, the difference is that with asyncio, it is the task of the developer rather
than the OS to decide when to switch to the next task.
"""

import asyncio
import hashlib
import logging
from io import BytesIO
from timeit import default_timer as timer
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import urlsplit

import aiohttp
from aiohttp import ClientResponseError
from aiolimiter import AsyncLimiter
from openpyxl import load_workbook
from tenacity import (
    retry,
    retry_if_exception,
    stop_after_attempt,
)
from tqdm.asyncio import tqdm_asyncio

from .tenacity_custom_wait import custom_wait
from .utilities import is_server_error

logger = logging.getLogger(__name__)


class Retrieval:
    """Retrieval class for downloading and hashing resources.

    Args:
        user_agent (str): User agent string to use when downloading
        url_ignore (Optional[str]): Parts of url to ignore for special xlsx handling
    """

    ignore_mimetypes = ["application/octet-stream", "application/binary"]
    mimetypes = {
        "json": ["application/json"],
        "geojson": ["application/json", "application/geo+json"],
        "shp": ["application/zip", "application/x-zip-compressed"],
        "csv": ["text/csv", "application/zip", "application/x-zip-compressed"],
        "xls": ["application/vnd.ms-excel"],
        "xlsx": [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ],
    }
    signatures = {
        "json": [b"[", b" [", b"{", b" {"],
        "geojson": [b"[", b" [", b"{", b" {"],
        "shp": [b"PK\x03\x04"],
        "xls": [b"\xd0\xcf\x11\xe0"],
        "xlsx": [b"PK\x03\x04"],
    }

    def __init__(
        self,
        user_agent: str,
        netlocs: Set[str],
        url_ignore: Optional[str] = None,
    ) -> None:
        self._user_agent = user_agent
        self._url_ignore: Optional[str] = url_ignore
        # Limit to 4 connections per second to a host
        self._rate_limiters = {
            netloc: AsyncLimiter(4, 1) for netloc in netlocs
        }

    @retry(
        reraise=True,
        retry=retry_if_exception(is_server_error),
        stop=stop_after_attempt(3),
        wait=custom_wait(multiplier=2, min=4),
    )
    async def fetch(
        self,
        url: str,
        resource_id: str,
        resource_format: str,
        session: aiohttp.ClientSession,
    ) -> Tuple:
        """Asynchronous code to download a resource and hash it with rate
        limiting and exception handling. Returns a tuple with resource
        information including hashes.

        Args:
            url (str): Resource to get
            resource_id (str): Resource id
            resource_format (str): Resource format
            session (Union[aiohttp.ClientSession, RateLimiter]): session to use for requests

        Returns:
            Tuple: Resource information including hash
        """
        async with session.get(
            url, allow_redirects=True, chunked=True
        ) as response:
            status = response.status
            if status != 200:
                exception = ClientResponseError(
                    code=status,
                    message=response.reason,
                    request_info=response.request_info,
                    history=response.history,
                )
                raise exception
            headers = response.headers
            http_size = headers.get("Content-Length")
            if http_size:
                http_size = int(http_size)
            last_modified = headers.get("Last-Modified")
            etag = headers.get("Etag")
            if etag:
                return resource_id, http_size, last_modified, etag, 200
            if http_size and int(http_size) > 419430400:
                return resource_id, http_size, last_modified, None, -11

            mimetype = headers.get("Content-Type")
            iterator = response.content.iter_any()
            first_chunk = await anext(iterator)
            size = len(first_chunk)
            signature = first_chunk[:4]
            if (
                resource_format == "xlsx"
                and (
                    mimetype == self.mimetypes["xlsx"][0]
                    or mimetype in self.ignore_mimetypes
                )
                and signature == self.signatures["xlsx"][0]
                and (self._url_ignore not in url if self._url_ignore else True)
            ):
                xlsxbuffer = bytearray(first_chunk)
                async for chunk in iterator:
                    size += len(chunk)
                    xlsxbuffer.extend(chunk)
                workbook = load_workbook(
                    filename=BytesIO(xlsxbuffer), read_only=True
                )
                md5hash = hashlib.md5()
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for cols in sheet.iter_rows(values_only=True):
                        md5hash.update(bytes(str(cols), "utf-8"))
                workbook.close()
                xlsxbuffer = None
            else:
                md5hash = hashlib.md5(first_chunk)
                async for chunk in iterator:
                    size += len(chunk)
                    md5hash.update(chunk)
            hash = md5hash.hexdigest()
            if mimetype not in self.ignore_mimetypes:
                expected_mimetypes = self.mimetypes.get(resource_format)
                if expected_mimetypes is not None:
                    if not any(x in mimetype for x in expected_mimetypes):
                        return (
                            resource_id,
                            size,
                            last_modified,
                            hash,
                            -1,
                        )
            expected_signatures = self.signatures.get(resource_format)
            if expected_signatures is not None:
                if not any(
                    signature[: len(x)] == x for x in expected_signatures
                ):
                    return resource_id, size, last_modified, hash, -2
            if size != http_size:
                return resource_id, size, last_modified, hash, -3

            return resource_id, size, last_modified, hash, 0

    async def process(
        self,
        metadata: Tuple,
        session: aiohttp.ClientSession,
    ) -> Tuple:
        """Asynchronous code to download a resource and hash it. Returns a tuple with
        resource information including hashes.

        Args:
            metadata (Tuple): Resource to be checked
            session (Union[aiohttp.ClientSession, RateLimiter]): session to use for requests

        Returns:
            Tuple: Resource information including hash
        """
        url = metadata[0]
        resource_id = metadata[1]
        resource_format = metadata[2]

        host = urlsplit(url).netloc

        async with self._rate_limiters[host]:
            try:
                return await self.fetch(
                    url, resource_id, resource_format, session
                )
            except ClientResponseError as ex:
                logger.error(f"{ex.status} {ex.message} {ex.request_info.url}")
                return resource_id, None, None, None, ex.status
            except Exception as ex:
                logger.error(ex)
                return resource_id, None, None, None, -101

    async def check_urls(
        self, resources_to_get: List[Tuple]
    ) -> Dict[str, Tuple]:
        """Asynchronous code to download resources and hash them. Return dictionary with
        resources information including hashes.

        Args:
            resources_to_get (List[Tuple]): List of resources to get
            loop (uvloop.Loop): Event loop to use

        Returns:
            Dict[str, Tuple]: Resources information including hashes
        """
        tasks = []

        # Maximum of 10 simultaneous connections to a host
        conn = aiohttp.TCPConnector(limit_per_host=10)
        # Can set some timeouts here if needed
        timeout = aiohttp.ClientTimeout(total=5 * 60, sock_connect=30)
        async with aiohttp.ClientSession(
            connector=conn,
            timeout=timeout,
            headers={"User-Agent": self._user_agent},
        ) as session:
            for metadata in resources_to_get:
                task = self.process(metadata, session)
                tasks.append(task)
            responses = {}
            for f in tqdm_asyncio.as_completed(tasks, total=len(tasks)):
                (
                    resource_id,
                    http_size,
                    http_last_modified,
                    hash,
                    status,
                ) = await f

                responses[resource_id] = (
                    http_size,
                    http_last_modified,
                    hash,
                    status,
                )
            return responses

    def retrieve(self, resources_to_get: List[Tuple]) -> Dict[str, Tuple]:
        """Download resources and hash them. Return dictionary with resources information
        including hashes.

        Args:
            resources_to_get (List[Tuple]): List of resources to get

        Returns:
            Dict[str, Tuple]: Resources information including hashes
        """

        start_time = timer()
        results = asyncio.run(self.check_urls(resources_to_get))
        logger.info(f"Execution time: {timer() - start_time} seconds")
        return results
